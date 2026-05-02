from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime
from io import BytesIO
import re
from database import get_db, log_action
import models
from dependencies import get_current_admin, require_coordinador

router = APIRouter(prefix="/api/admin", tags=["admin"])


# ── Me ────────────────────────────────────────────────────────

@router.get("/me")
def get_admin_me(
    admin: models.OpsAdminUser = Depends(get_current_admin),
):
    return {
        "id": admin.id,
        "username": admin.username,
        "full_name": admin.full_name,
        "role": admin.role,
    }


# ── Dashboard ─────────────────────────────────────────────────

@router.get("/dashboard-stats")
def get_dashboard_stats(
    period_type: str = None,
    year: int = None,
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    # Get current period
    current_period = db.query(models.DimPeriod).filter_by(is_current=True).first()

    # Total students
    total_students = db.query(models.OpsStudent).filter_by(is_active=True).count()

    # By service type (from enrollments)
    query_base = db.query(models.OpsStudentEnrollment)
    if current_period:
        query_base = query_base.filter_by(period_id=current_period.id)

    ss_count = query_base.filter_by(service_type="servicio_social").count()
    pp_count = query_base.filter_by(service_type="practica_profesional").count()

    # Pending uploads
    pending_uploads = db.query(models.OpsUploadStatus).filter_by(
        current_status="pending_review"
    ).count()

    # Pending requests
    pending_enrollments = db.query(models.FactEnrollment).filter_by(status="pending").count()
    pending_changes = db.query(models.FactChangeRequest).filter_by(status="pending").count()
    pending_requests = pending_enrollments + pending_changes

    # Full programs
    full_programs = db.query(models.OpsProgramAvailability).filter_by(is_full=True).count()

    # Recent activity (last 10 audit events)
    recent_activity = db.query(models.FactAuditLog).order_by(
        models.FactAuditLog.timestamp.desc()
    ).limit(10).all()

    # Top programs by demand
    top_programs = db.query(
        models.OpsProgramAvailability
    ).order_by(models.OpsProgramAvailability.used_slots.desc()).limit(5).all()

    top_programs_data = []
    for pa in top_programs:
        prog = db.query(models.DimProgram).filter_by(id=pa.program_id).first()
        if prog:
            top_programs_data.append({
                "folio": prog.folio,
                "name": prog.name,
                "used_slots": pa.used_slots,
                "max_slots": pa.max_slots,
                "is_full": pa.is_full,
            })

    return {
        "total_students": total_students,
        "ss_count": ss_count,
        "pp_count": pp_count,
        "pending_uploads": pending_uploads,
        "pending_requests": pending_requests,
        "full_programs": full_programs,
        "recent_activity": [
            {
                "id": a.id,
                "user_name": a.user_name,
                "action": a.action,
                "entity_type": a.entity_type,
                "timestamp": a.timestamp.isoformat() if a.timestamp else None,
            }
            for a in recent_activity
        ],
        "top_programs": top_programs_data,
    }


# ── Students ──────────────────────────────────────────────────

@router.get("/students")
def list_students(
    period_id: int = None,
    modality_code: str = None,
    service_type: str = None,
    search: str = None,
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    query = db.query(models.OpsStudent).filter_by(is_active=True)

    if search:
        query = query.filter(
            (models.OpsStudent.full_name.ilike(f"%{search}%")) |
            (models.OpsStudent.matricula.ilike(f"%{search}%"))
        )

    if modality_code:
        modality = db.query(models.DimModality).filter_by(code=modality_code).first()
        if modality:
            query = query.filter_by(modality_id=modality.id)

    students = query.order_by(models.OpsStudent.full_name).all()

    result = []
    for s in students:
        career = db.query(models.DimCareer).filter_by(id=s.career_id).first()
        modality = db.query(models.DimModality).filter_by(id=s.modality_id).first()

        # Current enrollment
        enrollment = db.query(models.OpsStudentEnrollment).filter_by(
            student_id=s.id
        ).order_by(models.OpsStudentEnrollment.enrolled_at.desc()).first()

        prog_name = None
        if enrollment and enrollment.program:
            prog_name = enrollment.program.name

        result.append({
            "id": s.id,
            "full_name": s.full_name,
            "matricula": s.matricula,
            "email": s.email,
            "career": {"code": career.code, "name": career.name} if career else None,
            "modality": {"code": modality.code, "name": modality.name} if modality else None,
            "enrollment_status": enrollment.status if enrollment else "not_enrolled",
            "service_type": enrollment.service_type if enrollment else None,
            "program_name": prog_name,
            "created_at": s.created_at.isoformat() if s.created_at else None,
        })
    return result


@router.get("/students/{student_id}")
def get_student_detail(
    student_id: int,
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    student = db.query(models.OpsStudent).filter_by(id=student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Alumno no encontrado.")

    career = db.query(models.DimCareer).filter_by(id=student.career_id).first()
    modality = db.query(models.DimModality).filter_by(id=student.modality_id).first()

    # Progress for all processes
    progress_records = db.query(models.OpsStudentProgress).filter_by(student_id=student_id).all()
    progress_data = []
    for p in progress_records:
        proc = db.query(models.DimProcessDefinition).filter_by(id=p.process_id).first()
        total_steps = db.query(models.DimProcessStep).filter_by(process_id=p.process_id).count()
        progress_data.append({
            "process_code": proc.code if proc else None,
            "process_name": proc.name if proc else None,
            "current_step": p.current_step,
            "total_steps": total_steps,
            "status": p.status,
            "started_at": p.started_at.isoformat() if p.started_at else None,
        })

    # Enrollments
    enrollments = db.query(models.OpsStudentEnrollment).filter_by(student_id=student_id).all()
    enrollment_data = []
    for e in enrollments:
        prog = db.query(models.DimProgram).filter_by(id=e.program_id).first()
        period = db.query(models.DimPeriod).filter_by(id=e.period_id).first()
        enrollment_data.append({
            "id": e.id,
            "status": e.status,
            "service_type": e.service_type,
            "program": {"folio": prog.folio, "name": prog.name, "dependency_name": prog.dependency_name} if prog else None,
            "period": period.name if period else None,
            "enrolled_at": e.enrolled_at.isoformat() if e.enrolled_at else None,
        })

    # Upload statuses
    upload_statuses = db.query(models.OpsUploadStatus).filter_by(student_id=student_id).all()
    uploads_data = []
    for us in upload_statuses:
        doc_type = db.query(models.DimDocumentType).filter_by(id=us.document_type_id).first()
        proc = db.query(models.DimProcessDefinition).filter_by(id=us.process_id).first()
        uploads_data.append({
            "document_type": {"code": doc_type.code, "name": doc_type.name} if doc_type else None,
            "process_code": proc.code if proc else None,
            "status": us.current_status,
            "attempt": us.current_attempt,
            "rejection_reason": us.last_rejection_reason,
        })

    return {
        "id": student.id,
        "full_name": student.full_name,
        "first_name": student.first_name,
        "last_name_paterno": student.last_name_paterno,
        "last_name_materno": student.last_name_materno,
        "matricula": student.matricula,
        "email": student.email,
        "career": {"code": career.code, "name": career.name} if career else None,
        "modality": {"code": modality.code, "name": modality.name} if modality else None,
        "is_active": student.is_active,
        "created_at": student.created_at.isoformat() if student.created_at else None,
        "last_login_at": student.last_login_at.isoformat() if student.last_login_at else None,
        "progress": progress_data,
        "enrollments": enrollment_data,
        "uploads": uploads_data,
    }


@router.post("/students/{student_id}/advance")
def advance_step(
    student_id: int,
    req: dict,
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    process_code = req.get("process_code", "").strip()
    notes = req.get("notes", "")

    proc = db.query(models.DimProcessDefinition).filter_by(code=process_code).first()
    if not proc:
        raise HTTPException(status_code=404, detail="Proceso no encontrado.")

    progress = db.query(models.OpsStudentProgress).filter_by(
        student_id=student_id, process_id=proc.id
    ).first()
    if not progress:
        raise HTTPException(status_code=404, detail="El alumno no tiene progreso en este proceso.")

    old_step = progress.current_step
    progress.current_step += 1
    progress.updated_at = datetime.utcnow()

    db.add(models.FactStepCompletion(
        student_id=student_id, process_id=proc.id,
        step_number=old_step, completed_by=f"admin:{admin.username}",
        notes=notes,
    ))

    log_action(db, "admin", admin.id, admin.full_name, "advance_step", "student", student_id,
               details_before={"step": old_step}, details_after={"step": progress.current_step})
    db.commit()
    return {"new_step": progress.current_step}


@router.post("/students/{student_id}/set-step")
def set_step(
    student_id: int,
    req: dict,
    admin: models.OpsAdminUser = Depends(require_coordinador),
    db: Session = Depends(get_db),
):
    process_code = req.get("process_code", "").strip()
    step = req.get("step", 1)

    proc = db.query(models.DimProcessDefinition).filter_by(code=process_code).first()
    if not proc:
        raise HTTPException(status_code=404, detail="Proceso no encontrado.")

    progress = db.query(models.OpsStudentProgress).filter_by(
        student_id=student_id, process_id=proc.id
    ).first()
    if not progress:
        progress = models.OpsStudentProgress(
            student_id=student_id, process_id=proc.id,
            current_step=step, status="active", started_at=datetime.utcnow(),
        )
        db.add(progress)
    else:
        old_step = progress.current_step
        progress.current_step = step
        progress.updated_at = datetime.utcnow()
        log_action(db, "admin", admin.id, admin.full_name, "advance_step", "student", student_id,
                   details_before={"step": old_step}, details_after={"step": step})

    db.commit()
    return {"step": step}


# ── Enrollments ───────────────────────────────────────────────

@router.get("/enrollments/pending")
def get_pending_enrollments(
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    enrollments = db.query(models.FactEnrollment).filter_by(status="pending").all()
    result = []
    for e in enrollments:
        student = db.query(models.OpsStudent).filter_by(id=e.student_id).first()
        prog = db.query(models.DimProgram).filter_by(id=e.program_id).first()
        avail = db.query(models.OpsProgramAvailability).filter_by(program_id=e.program_id).first()
        career = db.query(models.DimCareer).filter_by(id=student.career_id).first() if student else None
        result.append({
            "enrollment_id": e.id,
            "student": {
                "id": student.id,
                "full_name": student.full_name,
                "matricula": student.matricula,
                "career": career.name if career else None,
            } if student else None,
            "program": {
                "folio": prog.folio,
                "name": prog.name,
                "dependency_name": prog.dependency_name,
                "max_slots": avail.max_slots if avail else prog.max_slots,
                "used_slots": avail.used_slots if avail else 0,
            } if prog else None,
            "service_type": e.service_type,
            "enrolled_at": e.enrolled_at.isoformat() if e.enrolled_at else None,
        })
    return result


@router.get("/enrollments/{enrollment_id}")
def get_enrollment_detail(
    enrollment_id: int,
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    try:
        fact_e = db.query(models.FactEnrollment).filter_by(id=enrollment_id).first()
        if not fact_e:
            raise HTTPException(status_code=404, detail="Inscripción no encontrada.")

        student = db.query(models.OpsStudent).filter_by(id=fact_e.student_id).first()
        prog = db.query(models.DimProgram).filter_by(id=fact_e.program_id).first() if fact_e.program_id else None
        period = db.query(models.DimPeriod).filter_by(id=fact_e.period_id).first() if fact_e.period_id else None
        career = db.query(models.DimCareer).filter_by(id=student.career_id).first() if student and student.career_id else None
        avail = db.query(models.OpsProgramAvailability).filter_by(program_id=fact_e.program_id).first() if fact_e.program_id else None

        return {
            "id": fact_e.id,
            "status": fact_e.status,
            "service_type": fact_e.service_type,
            "folio_entered": fact_e.folio_entered,
            "enrolled_at": fact_e.enrolled_at.isoformat() if fact_e.enrolled_at else None,
            "validated_by": fact_e.validated_by,
            "validated_at": fact_e.validated_at.isoformat() if fact_e.validated_at else None,
            "rejection_reason": fact_e.rejection_reason,
            "student": {
                "id": student.id,
                "full_name": student.full_name,
                "matricula": student.matricula,
                "email": student.email,
                "career": {"code": career.code, "name": career.name} if career else None,
            } if student else None,
            "program": {
                "id": prog.id,
                "folio": prog.folio,
                "name": prog.name,
                "dependency_name": prog.dependency_name,
                "sector": prog.sector,
                "max_slots": avail.max_slots if avail else prog.max_slots,
                "used_slots": avail.used_slots if avail else 0,
                "available_slots": avail.available_slots if avail else prog.max_slots,
                "is_full": avail.is_full if avail else False,
            } if prog else None,
            "period": {
                "id": period.id,
                "code": period.code,
                "name": period.name,
            } if period else None,
        }

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error al obtener inscripción: {exc}")


@router.post("/enrollments/{enrollment_id}/approve")
def approve_enrollment(
    enrollment_id: int,
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    try:
        fact_e = db.query(models.FactEnrollment).filter_by(id=enrollment_id).first()
        if not fact_e:
            raise HTTPException(status_code=404, detail="Inscripción no encontrada.")

        fact_e.status = "approved"
        fact_e.validated_by = admin.username
        fact_e.validated_at = datetime.utcnow()

        # Find or create the ops enrollment record and set it to active
        ops_e = db.query(models.OpsStudentEnrollment).filter_by(
            student_id=fact_e.student_id, program_id=fact_e.program_id
        ).first()
        if ops_e:
            ops_e.status = "active"
        else:
            # Create it if it doesn't exist yet
            current_period = db.query(models.DimPeriod).filter_by(is_current=True).first()
            ops_e = models.OpsStudentEnrollment(
                student_id=fact_e.student_id,
                program_id=fact_e.program_id,
                period_id=fact_e.period_id or (current_period.id if current_period else None),
                service_type=fact_e.service_type,
                status="active",
            )
            db.add(ops_e)

        # Find or create OpsStudentProgress for the "inscripcion" process
        proc = db.query(models.DimProcessDefinition).filter_by(code="inscripcion").first()
        if proc:
            progress = db.query(models.OpsStudentProgress).filter_by(
                student_id=fact_e.student_id, process_id=proc.id
            ).first()
            if not progress:
                progress = models.OpsStudentProgress(
                    student_id=fact_e.student_id, process_id=proc.id,
                    current_step=3, status="active", started_at=datetime.utcnow(),
                )
                db.add(progress)
            else:
                progress.current_step = 3
                progress.status = "active"
                if not progress.started_at:
                    progress.started_at = datetime.utcnow()

            # Mark steps 1 and 2 as completed
            for step_num in [1, 2]:
                existing = db.query(models.FactStepCompletion).filter_by(
                    student_id=fact_e.student_id, process_id=proc.id, step_number=step_num
                ).first()
                if not existing:
                    db.add(models.FactStepCompletion(
                        student_id=fact_e.student_id, process_id=proc.id,
                        step_number=step_num, completed_by=f"admin:{admin.username}",
                    ))

        db.add(models.FactApprovalAction(
            entity_type="enrollment", entity_id=enrollment_id,
            action="approved", performed_by=admin.username,
        ))

        log_action(db, "admin", admin.id, admin.full_name, "approve_enrollment",
                   "enrollment", enrollment_id)
        db.commit()
        return {"message": "Inscripción aprobada. El alumno está en paso 3.", "enrollment_id": enrollment_id}

    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al aprobar inscripción: {exc}")


@router.post("/enrollments/{enrollment_id}/reject")
def reject_enrollment(
    enrollment_id: int,
    req: dict,
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    reason = req.get("reason", "").strip()
    fact_e = db.query(models.FactEnrollment).filter_by(id=enrollment_id).first()
    if not fact_e:
        raise HTTPException(status_code=404, detail="Inscripción no encontrada.")

    fact_e.status = "rejected"
    fact_e.rejection_reason = reason

    ops_e = db.query(models.OpsStudentEnrollment).filter_by(
        student_id=fact_e.student_id, program_id=fact_e.program_id
    ).first()
    if ops_e:
        ops_e.status = "cancelled"

    # Release slot
    avail = db.query(models.OpsProgramAvailability).filter_by(program_id=fact_e.program_id).first()
    if avail and avail.used_slots > 0:
        avail.used_slots -= 1
        avail.available_slots = avail.max_slots - avail.used_slots
        avail.is_full = False

    db.add(models.FactApprovalAction(
        entity_type="enrollment", entity_id=enrollment_id,
        action="rejected", performed_by=admin.username, reason=reason,
    ))

    log_action(db, "admin", admin.id, admin.full_name, "reject_enrollment",
               "enrollment", enrollment_id, details_after={"reason": reason})
    db.commit()
    return {"message": "Inscripción rechazada.", "enrollment_id": enrollment_id}


# ── Change Requests ───────────────────────────────────────────

@router.get("/change-requests/pending")
def get_pending_change_requests(
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    requests = db.query(models.FactChangeRequest).filter_by(status="pending").all()
    result = []
    for r in requests:
        student = db.query(models.OpsStudent).filter_by(id=r.student_id).first()
        current_prog = db.query(models.DimProgram).filter_by(id=r.current_program_id).first()
        new_prog = db.query(models.DimProgram).filter_by(id=r.new_program_id).first() if r.new_program_id else None
        result.append({
            "id": r.id,
            "request_type": r.request_type,
            "student": {
                "id": student.id,
                "full_name": student.full_name,
                "matricula": student.matricula,
            } if student else None,
            "current_program": {
                "folio": current_prog.folio,
                "name": current_prog.name,
                "dependency_name": current_prog.dependency_name,
            } if current_prog else None,
            "new_program": {
                "folio": new_prog.folio,
                "name": new_prog.name,
                "dependency_name": new_prog.dependency_name,
            } if new_prog else None,
            "justification": r.justification,
            "submitted_at": r.submitted_at.isoformat() if r.submitted_at else None,
        })
    return result


@router.get("/change-requests/{request_id}")
def get_change_request(
    request_id: int,
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    r = db.query(models.FactChangeRequest).filter_by(id=request_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada.")

    student = db.query(models.OpsStudent).filter_by(id=r.student_id).first()
    current_prog = db.query(models.DimProgram).filter_by(id=r.current_program_id).first()
    new_prog = db.query(models.DimProgram).filter_by(id=r.new_program_id).first() if r.new_program_id else None

    return {
        "id": r.id,
        "request_type": r.request_type,
        "status": r.status,
        "student": {"id": student.id, "full_name": student.full_name, "matricula": student.matricula} if student else None,
        "current_program": {"folio": current_prog.folio, "name": current_prog.name} if current_prog else None,
        "new_program": {"folio": new_prog.folio, "name": new_prog.name} if new_prog else None,
        "justification": r.justification,
        "submitted_at": r.submitted_at.isoformat() if r.submitted_at else None,
    }


@router.post("/change-requests/{request_id}/approve")
def approve_change_request(
    request_id: int,
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    r = db.query(models.FactChangeRequest).filter_by(id=request_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada.")

    r.status = "approved"

    enrollment = db.query(models.OpsStudentEnrollment).filter_by(
        student_id=r.student_id, status="active"
    ).first()

    if r.request_type == "cambio" and r.new_program_id and enrollment:
        old_prog_id = enrollment.program_id

        # Release old slot
        old_avail = db.query(models.OpsProgramAvailability).filter_by(program_id=old_prog_id).first()
        if old_avail and old_avail.used_slots > 0:
            old_avail.used_slots -= 1
            old_avail.available_slots = old_avail.max_slots - old_avail.used_slots
            old_avail.is_full = False

        # Take new slot
        new_avail = db.query(models.OpsProgramAvailability).filter_by(program_id=r.new_program_id).first()
        if new_avail:
            new_avail.used_slots += 1
            new_avail.available_slots = max(0, new_avail.max_slots - new_avail.used_slots)
            new_avail.is_full = new_avail.used_slots >= new_avail.max_slots

        # Update enrollment
        enrollment.program_id = r.new_program_id

    elif r.request_type == "baja" and enrollment:
        enrollment.status = "cancelled"
        old_avail = db.query(models.OpsProgramAvailability).filter_by(program_id=enrollment.program_id).first()
        if old_avail and old_avail.used_slots > 0:
            old_avail.used_slots -= 1
            old_avail.available_slots = old_avail.max_slots - old_avail.used_slots
            old_avail.is_full = False

    db.add(models.FactApprovalAction(
        entity_type="change_request", entity_id=request_id,
        action="approved", performed_by=admin.username,
    ))

    log_action(db, "admin", admin.id, admin.full_name, f"approve_{r.request_type}_request",
               "change_request", request_id)
    db.commit()
    return {"message": "Solicitud aprobada.", "request_id": request_id}


@router.post("/change-requests/{request_id}/reject")
def reject_change_request(
    request_id: int,
    req: dict,
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    reason = req.get("reason", "").strip()
    r = db.query(models.FactChangeRequest).filter_by(id=request_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada.")

    r.status = "rejected"

    db.add(models.FactApprovalAction(
        entity_type="change_request", entity_id=request_id,
        action="rejected", performed_by=admin.username, reason=reason,
    ))

    log_action(db, "admin", admin.id, admin.full_name, f"reject_{r.request_type}_request",
               "change_request", request_id, details_after={"reason": reason})
    db.commit()
    return {"message": "Solicitud rechazada.", "request_id": request_id}


# ── Config ────────────────────────────────────────────────────

@router.get("/config")
def get_config(
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    configs = db.query(models.DimSystemConfig).all()
    return [
        {
            "config_key": c.config_key,
            "config_value": c.config_value,
            # aliases for frontend compatibility
            "key": c.config_key,
            "value": c.config_value,
            "config_type": c.config_type,
            "description": c.description,
            "updated_at": c.updated_at.isoformat() if c.updated_at else None,
            "updated_by": c.updated_by,
        }
        for c in configs
    ]


@router.put("/config/{config_key}")
def update_config(
    config_key: str,
    req: dict,
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    config = db.query(models.DimSystemConfig).filter_by(config_key=config_key).first()
    if not config:
        raise HTTPException(status_code=404, detail="Clave de configuración no encontrada.")

    old_value = config.config_value
    new_value = req.get("value", "").strip()
    config.config_value = new_value
    config.updated_at = datetime.utcnow()
    config.updated_by = admin.username

    log_action(db, "admin", admin.id, admin.full_name, "update_config",
               "system_config", config.id,
               details_before={"key": config_key, "value": old_value},
               details_after={"key": config_key, "value": new_value})
    db.commit()
    return {"config_key": config_key, "config_value": new_value}



# ── Audit Log ─────────────────────────────────────────────────

@router.get("/audit-log")
def get_audit_log(
    action: str = None,
    limit: int = 50,
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    import json as _json
    query = db.query(models.FactAuditLog)
    if action:
        query = query.filter(models.FactAuditLog.action == action)
    logs = query.order_by(models.FactAuditLog.timestamp.desc()).limit(limit).all()
    return [
        {
            "id": l.id,
            "user_type": l.user_type,
            "user_id": l.user_id,
            "user_name": l.user_name,
            "action": l.action,
            "entity_type": l.entity_type,
            "entity_id": l.entity_id,
            "details_before": _json.loads(l.details_before) if l.details_before else None,
            "details_after": _json.loads(l.details_after) if l.details_after else None,
            "created_at": l.timestamp.isoformat() if l.timestamp else None,
        }
        for l in logs
    ]


# ── Periods ───────────────────────────────────────────────────

@router.get("/periods")
def list_periods(
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    periods = db.query(models.DimPeriod).order_by(models.DimPeriod.year.desc()).all()
    return [
        {
            "id": p.id, "code": p.code, "name": p.name,
            "period_type": p.period_type, "year": p.year,
            "start_date": str(p.start_date) if p.start_date else None,
            "end_date": str(p.end_date) if p.end_date else None,
            "enrollment_start": str(p.enrollment_start) if p.enrollment_start else None,
            "enrollment_end": str(p.enrollment_end) if p.enrollment_end else None,
            "is_current": p.is_current,
        }
        for p in periods
    ]


@router.post("/periods")
def create_period(
    req: dict,
    admin: models.OpsAdminUser = Depends(require_coordinador),
    db: Session = Depends(get_db),
):
    from datetime import date
    period = models.DimPeriod(
        code=req.get("code"),
        name=req.get("name"),
        period_type=req.get("period_type"),
        year=req.get("year"),
        start_date=date.fromisoformat(req["start_date"]) if req.get("start_date") else None,
        end_date=date.fromisoformat(req["end_date"]) if req.get("end_date") else None,
        enrollment_start=date.fromisoformat(req["enrollment_start"]) if req.get("enrollment_start") else None,
        enrollment_end=date.fromisoformat(req["enrollment_end"]) if req.get("enrollment_end") else None,
    )
    db.add(period)
    db.commit()
    db.refresh(period)
    return {"id": period.id, "name": period.name}


@router.put("/periods/{period_id}")
def update_period(
    period_id: int,
    req: dict,
    admin: models.OpsAdminUser = Depends(require_coordinador),
    db: Session = Depends(get_db),
):
    from datetime import date
    period = db.query(models.DimPeriod).filter_by(id=period_id).first()
    if not period:
        raise HTTPException(status_code=404, detail="Periodo no encontrado.")

    for field in ["code", "name", "period_type", "year"]:
        if field in req:
            setattr(period, field, req[field])
    for date_field in ["start_date", "end_date", "enrollment_start", "enrollment_end"]:
        if date_field in req and req[date_field]:
            setattr(period, date_field, date.fromisoformat(req[date_field]))

    db.commit()
    return {"message": "Periodo actualizado."}


@router.post("/periods/{period_id}/activate")
def activate_period(
    period_id: int,
    admin: models.OpsAdminUser = Depends(require_coordinador),
    db: Session = Depends(get_db),
):
    # Deactivate all
    db.query(models.DimPeriod).update({"is_current": False})
    # Activate this one
    period = db.query(models.DimPeriod).filter_by(id=period_id).first()
    if not period:
        raise HTTPException(status_code=404, detail="Periodo no encontrado.")
    period.is_current = True

    # Update system config
    cfg = db.query(models.DimSystemConfig).filter_by(config_key="current_period_id").first()
    if cfg:
        cfg.config_value = str(period_id)
    db.commit()
    return {"message": f"Periodo '{period.name}' activado."}


# ── Programs ──────────────────────────────────────────────────

@router.get("/careers")
def list_careers(
    db: Session = Depends(get_db),
    admin: models.OpsAdminUser = Depends(get_current_admin),
):
    careers = db.query(models.DimCareer).filter_by(is_active=True).order_by(models.DimCareer.name).all()
    return [{"id": c.id, "code": c.code, "name": c.name} for c in careers]


@router.get("/programs")
def list_programs(
    period_id: int = None,
    career_code: str = None,
    program_type: str = None,
    available_only: bool = False,
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    query = db.query(models.DimProgram).filter_by(is_active=True)

    if period_id:
        query = query.filter_by(period_id=period_id)
    if career_code:
        career = db.query(models.DimCareer).filter_by(code=career_code).first()
        if career:
            query = query.filter_by(career_id=career.id)
    if program_type:
        query = query.filter_by(program_type=program_type)
    if available_only:
        query = query.filter(models.DimProgram.status != "full")

    programs = query.order_by(models.DimProgram.folio).limit(3000).all()
    result = []
    for prog in programs:
        avail = db.query(models.OpsProgramAvailability).filter_by(program_id=prog.id).first()
        career = db.query(models.DimCareer).filter_by(id=prog.career_id).first()
        result.append({
            "id": prog.id,
            "folio": prog.folio,
            "name": prog.name,
            "program_type": prog.program_type,
            "career": {"code": career.code, "name": career.name} if career else None,
            "career_code": career.code if career else None,
            "career_name": career.code if career else None,  # show code (e.g. "LAE") as short label
            "dependency_name": prog.dependency_name,
            "sector": prog.sector,
            "max_slots": prog.max_slots,
            "used_slots": avail.used_slots if avail else 0,
            "available_slots": avail.available_slots if avail else prog.max_slots,
            "is_full": avail.is_full if avail else False,
            "status": prog.status,
        })
    return result


@router.post("/programs/upload-excel", response_model=None)
async def upload_programs_excel(
    file: UploadFile = File(...),
    admin: models.OpsAdminUser = Depends(require_coordinador),
    db: Session = Depends(get_db),
):
    """
    Upload an Excel file to upsert programs (dim_programs + ops_program_availability).

    Row 1 = "Exported data" title, Row 2 = headers, Row 3+ = data.
    Column order (0-indexed): 0=Evaluó, 1=Estado, 2=Folio, 3=Programa, 4=Tipo,
    5=Perfil, 6=. (cupo máximo), 7=Industria, 8=Sector, 9=Inscritos, 10=Cupo
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no está instalado. Agrega openpyxl a requirements.txt.")

    # Read file bytes
    file_content = await file.read()
    if not file_content:
        raise HTTPException(status_code=400, detail="El archivo está vacío.")

    try:
        wb = load_workbook(filename=BytesIO(file_content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo abrir el archivo Excel: {exc}")

    ws = wb.active

    # Map Tipo → program_type
    def map_tipo(tipo_raw):
        if tipo_raw is None:
            return None
        t = tipo_raw.lower().strip()
        if "pr" in t and "ctica" in t:
            return "practica_profesional"
        if "práctica" in t or "practica" in t:
            return "practica_profesional"
        if "servicio" in t:
            return "servicio_social"
        return None

    total_rows = 0
    new_count = 0
    updated_count = 0
    errors = []

    try:
        # Row 1 = title, Row 2 = headers, Row 3+ = data
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            # Skip completely empty rows
            if all(v is None for v in row):
                continue

            total_rows += 1

            try:
                # --- Extract fields by fixed column index ---
                evaluator_name = str(row[0]).strip() if row[0] else None
                # row[1] = Estado (not used for logic)
                folio = str(row[2]).strip() if row[2] else None
                if not folio:
                    errors.append({"row": row_idx, "reason": "Folio vacío o ausente."})
                    continue

                programa_name = str(row[3]).strip() if row[3] else None
                if not programa_name:
                    errors.append({"row": row_idx, "reason": f"Fila {row_idx}: Nombre de programa vacío."})
                    continue

                tipo_raw = str(row[4]).strip() if row[4] else None
                program_type = map_tipo(tipo_raw)
                if not program_type:
                    errors.append({"row": row_idx, "reason": f"Fila {row_idx} (folio {folio}): Tipo desconocido '{tipo_raw}'."})
                    continue

                profile = str(row[5]).strip() if row[5] else None

                max_slots = int(row[6]) if row[6] else 0

                dependency_name = str(row[7]).strip() if row[7] else None
                sector = str(row[8]).strip() if row[8] else None

                # Inscritos = used_slots (only applied on INSERT)
                inscritos = int(row[9]) if row[9] else 0

                # row[10] = Cupo (availability status text, ignored for logic)

                # --- Career matching via regex on parenthesised code ---
                career = None
                match = re.search(r'\(([A-Z]+)\)', str(profile or ''))
                career_code = match.group(1) if match else None
                career = db.query(models.DimCareer).filter_by(code=career_code).first() if career_code else None
                if not career:
                    errors.append({"row": row_idx, "reason": f"Fila {row_idx} (folio {folio}): Carrera no encontrada para perfil '{profile}'."})
                    continue

                # --- Upsert logic ---
                existing = db.query(models.DimProgram).filter(
                    models.DimProgram.folio == folio
                ).first()

                if existing:
                    # UPDATE — never reset used_slots
                    existing.name = programa_name
                    existing.program_type = program_type
                    existing.max_slots = max_slots
                    existing.dependency_name = dependency_name
                    existing.sector = sector
                    existing.evaluator_name = evaluator_name
                    # Update availability max_slots but preserve used_slots
                    avail = db.query(models.OpsProgramAvailability).filter_by(
                        program_id=existing.id
                    ).first()
                    if avail:
                        avail.max_slots = max_slots
                        avail.available_slots = max(0, max_slots - avail.used_slots)
                        avail.is_full = avail.used_slots >= max_slots
                    updated_count += 1
                else:
                    # INSERT new program
                    current_period = db.query(models.DimPeriod).filter_by(is_current=True).first()
                    period_id = current_period.id if current_period else None

                    new_prog = models.DimProgram(
                        folio=folio,
                        name=programa_name,
                        program_type=program_type,
                        career_id=career.id,
                        max_slots=max_slots,
                        dependency_name=dependency_name,
                        sector=sector,
                        evaluator_name=evaluator_name,
                        period_id=period_id,
                        status="active",
                        is_active=True,
                    )
                    db.add(new_prog)
                    db.flush()  # get new_prog.id

                    avail_slots = max(0, max_slots - inscritos)
                    new_avail = models.OpsProgramAvailability(
                        program_id=new_prog.id,
                        max_slots=max_slots,
                        used_slots=inscritos,
                        available_slots=avail_slots,
                        is_full=(inscritos >= max_slots),
                    )
                    db.add(new_avail)
                    new_count += 1

            except Exception as row_exc:
                errors.append({"row": row_idx, "reason": f"Error inesperado: {row_exc}"})
                continue

        summary = {
            "success": True,
            "total_rows": total_rows,
            "new_programs": new_count,
            "updated_programs": updated_count,
            "errors_count": len(errors),
            "errors": errors[:20],
            # aliases for frontend
            "processed": total_rows,
            "created": new_count,
            "updated": updated_count,
        }

        log_action(
            db, "admin", admin.id, admin.full_name,
            "upload_programs_excel", "dim_programs", None,
            details_after=summary,
        )
        db.commit()
        return summary

    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error procesando el archivo: {exc}")


@router.get("/programs/stats")
def get_program_stats(
    admin: models.OpsAdminUser = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    total = db.query(models.DimProgram).filter_by(is_active=True).count()
    ss_count = db.query(models.DimProgram).filter_by(program_type="servicio_social", is_active=True).count()
    pp_count = db.query(models.DimProgram).filter_by(program_type="practica_profesional", is_active=True).count()
    full_count = db.query(models.OpsProgramAvailability).filter_by(is_full=True).count()
    return {
        "total": total,
        "servicio_social": ss_count,
        "practica_profesional": pp_count,
        "full": full_count,
        "available": total - full_count,
    }


# ── Users ─────────────────────────────────────────────────────

@router.get("/users")
def list_users(
    admin: models.OpsAdminUser = Depends(require_coordinador),
    db: Session = Depends(get_db),
):
    users = db.query(models.OpsAdminUser).all()
    return [
        {
            "id": u.id, "username": u.username, "full_name": u.full_name,
            "role": u.role, "is_active": u.is_active,
            "created_at": u.created_at.isoformat() if u.created_at else None,
        }
        for u in users
    ]
